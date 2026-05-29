"""
ingest_cognition_drive.py

Purpose:
- Crawl the COGNITION Google Drive folder.
- Read Google Docs hot takes.
- Read XLSX survey workbooks.
- Pair Google Docs and chart PNGs by question number, e.g. Q5 <-> Q5_chart.png.
- Create embeddings.
- Store records in a separate LanceDB table: cognition_insights.

Expected environment variables:
- GOOGLE_SERVICE_ACCOUNT_JSON
- OPENAI_API_KEY
- COGNITION_DRIVE_FOLDER_ID    optional; defaults to the folder tested in shell
- LANCEDB_PATH                 optional; defaults to /data/lancedb
- COGNITION_TABLE_NAME         optional; defaults to cognition_insights
"""

from __future__ import annotations

import io
import json
import os
import re
import tempfile
import time
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional, Tuple

import lancedb
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openai import OpenAI

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


ROOT_FOLDER_ID = os.getenv("COGNITION_DRIVE_FOLDER_ID", "1asg9N24MasA3vn16OJXEwTsdqVoRwRyu")
LANCEDB_PATH = os.getenv("LANCEDB_PATH", "/data/lancedb")
TABLE_NAME = os.getenv("COGNITION_TABLE_NAME", "cognition_insights")
EMBED_MODEL = os.getenv("COGNITION_EMBED_MODEL", "text-embedding-3-small")

GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/documents.readonly",
]


DOC_MIME = "application/vnd.google-apps.document"
FOLDER_MIME = "application/vnd.google-apps.folder"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PNG_MIME = "image/png"
JPEG_MIMES = {"image/jpeg", "image/jpg"}


@dataclass
class CognitionInsight:
    id: str
    source_type: str
    folder_id: str
    folder_name: str
    question_id: str
    title: str
    headline: str
    summary: str
    body: str
    doc_id: str
    doc_name: str
    chart_image_id: str
    chart_image_name: str
    workbook_id: str
    workbook_name: str
    content_kind: str
    drive_url: str
    indexed_at_utc: str
    vector: List[float]


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u2014", "—")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def extract_question_id(name: str) -> str:
    match = re.search(r"\bQ(\d{1,2})\b|^Q(\d{1,2})_", name, flags=re.IGNORECASE)
    if not match:
        return ""
    num = match.group(1) or match.group(2)
    return f"Q{int(num)}"


def slugish(value: str) -> str:
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value).strip("-")
    return value[:120] or "untitled"


def get_google_services():
    raw_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not raw_json:
        raise RuntimeError("Missing GOOGLE_SERVICE_ACCOUNT_JSON environment variable.")

    creds = service_account.Credentials.from_service_account_info(
        json.loads(raw_json),
        scopes=GOOGLE_SCOPES,
    )

    drive = build("drive", "v3", credentials=creds)
    docs = build("docs", "v1", credentials=creds)
    return drive, docs


def list_children(drive, folder_id: str, page_size: int = 1000) -> List[Dict[str, Any]]:
    files: List[Dict[str, Any]] = []
    token = None

    while True:
        response = drive.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            pageSize=page_size,
            pageToken=token,
            fields="nextPageToken, files(id,name,mimeType,modifiedTime,size)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()

        files.extend(response.get("files", []))
        token = response.get("nextPageToken")
        if not token:
            break

    return files


def read_google_doc_text(docs, doc_id: str) -> str:
    doc = docs.documents().get(documentId=doc_id).execute()
    parts: List[str] = []

    for element in doc.get("body", {}).get("content", []):
        para = element.get("paragraph")
        if not para:
            continue

        line = ""
        for e in para.get("elements", []):
            text_run = e.get("textRun")
            if text_run:
                line += text_run.get("content", "")

        line = line.strip()
        if line:
            parts.append(line)

    return "\n".join(parts).strip()


def download_drive_file(drive, file_id: str) -> bytes:
    request = drive.files().get_media(fileId=file_id)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)

    done = False
    while not done:
        _, done = downloader.next_chunk()

    return buffer.getvalue()


def read_xlsx_records(drive, workbook_file: Dict[str, Any], folder_name: str) -> List[Dict[str, str]]:
    if load_workbook is None:
        print("WARNING: openpyxl is not installed. Skipping XLSX extraction.")
        return []

    file_id = workbook_file["id"]
    content = download_drive_file(drive, file_id)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    wb = load_workbook(tmp_path, data_only=True)
    records: List[Dict[str, str]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: List[str] = []

        for row in ws.iter_rows(values_only=True):
            values = [normalize_text(v) for v in row if normalize_text(v)]
            if values:
                rows.append(" | ".join(values))

        if not rows:
            continue

        question_id = extract_question_id(sheet_name)
        text = "\n".join(rows)
        title = rows[0] if rows else folder_name
        question = rows[1] if len(rows) > 1 else sheet_name

        records.append({
            "question_id": question_id,
            "title": title,
            "headline": question,
            "body": text,
            "summary": first_meaningful_summary(text),
            "workbook_id": workbook_file["id"],
            "workbook_name": workbook_file["name"],
        })

    return records


def first_meaningful_summary(text: str, max_chars: int = 550) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return ""

    # Prefer quoted hot-take summary if present.
    for line in lines:
        if len(line) > 80 and not line.startswith("[") and "COGNITION Weekly Hot Take" not in line:
            return line[:max_chars].strip()

    return " ".join(lines[:3])[:max_chars].strip()


def extract_headline(text: str, fallback_name: str) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for line in lines:
        clean = line.strip("*").strip()
        if clean and clean.lower() != "cognition weekly hot take":
            if "· COGNITION" in clean:
                continue
            if clean.startswith("[") and clean.endswith("]"):
                continue
            return clean
    return fallback_name


def build_chart_lookup(files: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    lookup: Dict[str, Dict[str, Any]] = {}

    for f in files:
        mime = f.get("mimeType", "")
        if mime != PNG_MIME and mime not in JPEG_MIMES:
            continue

        qid = extract_question_id(f.get("name", ""))
        if qid and qid not in lookup:
            lookup[qid] = f

    return lookup


def make_embedding(client: OpenAI, text: str, retries: int = 4) -> List[float]:
    text = text[:12000]

    for attempt in range(retries):
        try:
            response = client.embeddings.create(
                model=EMBED_MODEL,
                input=text,
            )
            return response.data[0].embedding
        except Exception as exc:
            if attempt == retries - 1:
                raise
            sleep_for = 2 ** attempt
            print(f"Embedding failed: {exc}. Retrying in {sleep_for}s...")
            time.sleep(sleep_for)

    raise RuntimeError("Embedding failed unexpectedly.")


def doc_drive_url(doc_id: str) -> str:
    return f"https://docs.google.com/document/d/{doc_id}/edit"


def image_proxy_url(file_id: str) -> str:
    return f"/api/cognition/image/{file_id}"


def build_records_from_folder(
    drive,
    docs,
    client: OpenAI,
    folder: Dict[str, Any],
) -> List[CognitionInsight]:
    folder_id = folder["id"]
    folder_name = folder["name"]

    files = list_children(drive, folder_id)
    chart_by_qid = build_chart_lookup(files)

    docs_files = [f for f in files if f.get("mimeType") == DOC_MIME]
    workbook_files = [f for f in files if f.get("mimeType") == XLSX_MIME]

    records: List[CognitionInsight] = []

    # Pattern B: Google Docs + PNG charts.
    for doc_file in docs_files:
        doc_id = doc_file["id"]
        doc_name = doc_file["name"]
        qid = extract_question_id(doc_name)

        try:
            body = read_google_doc_text(docs, doc_id)
        except Exception as exc:
            print(f"WARNING: Could not read doc {doc_name}: {exc}")
            continue

        if not body:
            continue

        chart = chart_by_qid.get(qid, {})
        headline = extract_headline(body, doc_name)
        summary = first_meaningful_summary(body)

        searchable_text = "\n".join([
            folder_name,
            qid,
            doc_name,
            headline,
            summary,
            body,
        ])

        vector = make_embedding(client, searchable_text)

        rec_id = f"doc-{doc_id}"
        records.append(CognitionInsight(
            id=rec_id,
            source_type="cognition_insight",
            folder_id=folder_id,
            folder_name=folder_name,
            question_id=qid,
            title=headline,
            headline=headline,
            summary=summary,
            body=body,
            doc_id=doc_id,
            doc_name=doc_name,
            chart_image_id=chart.get("id", ""),
            chart_image_name=chart.get("name", ""),
            workbook_id="",
            workbook_name="",
            content_kind="google_doc_hot_take",
            drive_url=doc_drive_url(doc_id),
            indexed_at_utc=now_utc(),
            vector=vector,
        ))

    # Pattern A: XLSX workbook + PNG charts.
    for workbook_file in workbook_files:
        try:
            workbook_records = read_xlsx_records(drive, workbook_file, folder_name)
        except Exception as exc:
            print(f"WARNING: Could not read workbook {workbook_file.get('name')}: {exc}")
            continue

        for item in workbook_records:
            qid = item.get("question_id", "")
            chart = chart_by_qid.get(qid, {})

            body = item.get("body", "")
            headline = item.get("headline") or item.get("title") or folder_name
            summary = item.get("summary") or first_meaningful_summary(body)

            searchable_text = "\n".join([
                folder_name,
                qid,
                headline,
                summary,
                body,
            ])

            vector = make_embedding(client, searchable_text)

            rec_id = f"xlsx-{workbook_file['id']}-{qid or slugish(headline)}"
            records.append(CognitionInsight(
                id=rec_id,
                source_type="cognition_insight",
                folder_id=folder_id,
                folder_name=folder_name,
                question_id=qid,
                title=headline,
                headline=headline,
                summary=summary,
                body=body,
                doc_id="",
                doc_name="",
                chart_image_id=chart.get("id", ""),
                chart_image_name=chart.get("name", ""),
                workbook_id=workbook_file["id"],
                workbook_name=workbook_file["name"],
                content_kind="xlsx_survey_data",
                drive_url=f"https://drive.google.com/file/d/{workbook_file['id']}/view",
                indexed_at_utc=now_utc(),
                vector=vector,
            ))

    return records


def create_or_replace_table(records: List[CognitionInsight]) -> None:
    if not records:
        print("No records to write.")
        return

    db = lancedb.connect(LANCEDB_PATH)
    rows = [asdict(r) for r in records]

    if TABLE_NAME in db.table_names():
        db.drop_table(TABLE_NAME)

    db.create_table(TABLE_NAME, data=rows)
    print(f"Created LanceDB table '{TABLE_NAME}' with {len(rows)} records at {LANCEDB_PATH}")


def main() -> None:
    print("Starting COGNITION Drive ingestion...")
    print(f"Root folder: {ROOT_FOLDER_ID}")
    print(f"LanceDB path: {LANCEDB_PATH}")
    print(f"Table: {TABLE_NAME}")

    drive, docs = get_google_services()
    client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))

    root_children = list_children(drive, ROOT_FOLDER_ID)
    folders = [f for f in root_children if f.get("mimeType") == FOLDER_MIME]

    print(f"Found {len(folders)} top-level Cognition folders.")

    all_records: List[CognitionInsight] = []

    for i, folder in enumerate(folders, 1):
        print(f"\n[{i}/{len(folders)}] Scanning: {folder['name']}")
        try:
            folder_records = build_records_from_folder(drive, docs, client, folder)
            print(f"  Indexed candidate records: {len(folder_records)}")
            all_records.extend(folder_records)
        except Exception as exc:
            print(f"  ERROR scanning folder {folder['name']}: {exc}")

    create_or_replace_table(all_records)

    print("\nDone.")
    print(f"Total Cognition insight records indexed: {len(all_records)}")


if __name__ == "__main__":
    main()
